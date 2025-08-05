import streamlit as st
import io
from openpyxl import load_workbook, Workbook
from copy import copy

st.set_page_config(page_title="Exportar Clubes com Movimentação", layout="centered")
st.title("♣️ Exportar Abas de Clubes com Movimentação")

uploaded_file = st.file_uploader("Envie sua planilha Excel (.xlsx)", type="xlsx")

if uploaded_file:
    sufixo = st.text_input("Data do Fechamento", value="exportado")

    if sufixo:
        # ✅ Carrega planilha com valores (sem fórmulas)
        wb_original = load_workbook(uploaded_file, data_only=True)

        # ✅ Ler aba "FECHAMENTO INTERNO" para identificar clubes com movimentação
        try:
            aba_controle = wb_original["FECHAMENTO INTERNO"]
            clubes_com_movimentacao = []

            for row in range(3, 23):  # A3:A22 e G3:G22
                nome = aba_controle[f"A{row}"].value
                valor = aba_controle[f"G{row}"].value

                if nome and isinstance(valor, (int, float)) and valor != 0:
                    clubes_com_movimentacao.append(nome)

        except KeyError:
            st.error("A aba 'FECHAMENTO INTERNO' não foi encontrada na planilha.")
            st.stop()

        if not clubes_com_movimentacao:
            st.warning("Nenhum clube com movimentação foi encontrado.")
            st.stop()

        # ✅ Exibir somente essas abas
        st.subheader("Selecionar clubes com movimentação:")
        selected_sheets = []
        for nome_aba in clubes_com_movimentacao:
            if nome_aba in wb_original.sheetnames:
                if st.checkbox(nome_aba, value=FALSE):
                    selected_sheets.append(nome_aba)

        if selected_sheets:
            st.subheader("Baixar arquivos exportados:")
            for sheet_name in selected_sheets:
                original_sheet = wb_original[sheet_name]
                new_wb = Workbook()
                new_sheet = new_wb.active
                new_sheet.title = sheet_name

                # Copiar dados + estilos
                for row in original_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                # Tamanho de colunas e linhas
                for col_letter, dim in original_sheet.column_dimensions.items():
                    new_sheet.column_dimensions[col_letter].width = dim.width
                for row_num, dim in original_sheet.row_dimensions.items():
                    new_sheet.row_dimensions[row_num].height = dim.height

                # Mesclagens
                for merged_range in original_sheet.merged_cells.ranges:
                    new_sheet.merge_cells(str(merged_range))

                # Salvar como download
                output = io.BytesIO()
                new_wb.save(output)
                output.seek(0)

                filename = f"{sheet_name} {sufixo}.xlsx"
                st.download_button(
                    label=f"⬇️ Baixar: {filename}",
                    data=output,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("Selecione ao menos um clube.")

